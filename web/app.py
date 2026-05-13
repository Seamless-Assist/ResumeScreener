"""SeamlessAssist Candidate Finder web UI backed by live jobs + saved role runs."""
from __future__ import annotations

# Load .env from the project root before any other imports read os.getenv()
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(dotenv_path=str(__import__('pathlib').Path(__file__).resolve().parent.parent / '.env'))
except ImportError:
    pass

import json
import os
import re
import subprocess
import sys
import threading
import time
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import urllib.parse as _urlparse
import urllib.request as _urlrequest
from flask import Flask, Response, abort, jsonify, redirect, has_request_context, render_template, request, session, url_for

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from sa_candidate_finder.manatal_jobs import fetch_all_jobs
from sa_candidate_finder.manatal_candidates import is_excluded_stage, load_role_stage_snapshot, sync_role_stages, invalidate_stage_snapshot, update_stage_snapshot_entry
from sa_candidate_finder.secrets import MANATAL_API_KEY, OPENAI_API_KEY, MANATAL_BASE_URL, GOODFIT_API_KEY
from openai import OpenAI as _OpenAI

_openai_client = _OpenAI(api_key=OPENAI_API_KEY)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SA_WEB_SESSION_SECRET", "sa-dev-session-secret")


@app.template_filter("to_pacific")
def _to_pacific_filter(value: str) -> str:
    """Convert a UTC ISO-8601 string to a Pacific time display string (PST/PDT)."""
    dt = _parse_iso_utc(value)
    if dt is None:
        return value or "-"
    from datetime import timedelta
    # Determine Pacific offset: PDT (UTC-7) Mar–Nov, PST (UTC-8) otherwise
    # Use a simple DST approximation based on the month
    month = dt.month
    is_dst = 3 <= month <= 11
    offset = timedelta(hours=-7 if is_dst else -8)
    label = "PDT" if is_dst else "PST"
    pt = dt + offset
    return pt.strftime(f"%-m/%-d/%Y %-I:%M %p {label}")

RESULTS_DIR = ROOT / "results"
SESSION_RESULTS_ROOT = ROOT / "cache" / "session_results"

RERANK_JOBS: dict[str, dict] = {}
RERANK_LOCK = threading.Lock()

STATUS_META: dict[str, dict] = {
    "Sourcing": {"color": "gold", "hex": "#A67D44"},
    "Shortlist review": {"color": "taupe", "hex": "#CDBCAB"},
    "Active outreach": {"color": "burgundy", "hex": "#5D1C34"},
    "Interviewing": {"color": "green", "hex": "#899481"},
    "On hold": {"color": "taupe", "hex": "#CDBCAB"},
    "Closed": {"color": "charcoal", "hex": "#11100F"},
}


def _normalize_role(name: str) -> str:
    token = (name or "").lower().strip()
    token = re.sub(
        r"[\s\-\u2013]+(?:t\d|tier\s*\d|remote|us|uk|ph|au|part.time|full.time|hourly)[^\w]*$",
        "",
        token,
    )
    token = re.sub(r"[^a-z0-9 ]+", " ", token)
    token = re.sub(r"\s+", " ", token).strip()
    return token


def _role_id_from_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _normalize_role(name)).strip("_")


def _get_session_id(create: bool = False) -> Optional[str]:
    if not has_request_context():
        return None
    sid = str(session.get("sa_session_id", "")).strip()
    if not sid and create:
        sid = uuid.uuid4().hex
        session["sa_session_id"] = sid
    return sid or None


def _session_results_dir(session_id: str) -> Path:
    return SESSION_RESULTS_ROOT / session_id


def _role_results_path(role_id: str, session_id: Optional[str] = None) -> Path:
    base = _session_results_dir(session_id) if session_id else RESULTS_DIR
    return base / f"role_{role_id}.json"


def _load_role_results(role_id: str) -> Optional[dict]:
    sid = _get_session_id(create=False)
    if sid:
        session_path = _role_results_path(role_id, sid)
        if session_path.exists():
            try:
                return json.loads(session_path.read_text(encoding="utf-8"))
            except Exception:
                pass

    path = _role_results_path(role_id)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _jobs_cache_fetched_at() -> Optional[datetime]:
    path = ROOT / "cache" / "jobs" / "all_jobs.json"
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    fetched_at = payload.get("fetched_at") if isinstance(payload, dict) else None
    if fetched_at is None:
        return None
    try:
        return datetime.fromtimestamp(float(fetched_at), tz=timezone.utc)
    except Exception:
        return None


def _parse_iso_utc(value: str) -> Optional[datetime]:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception:
        return None


def _parse_comma_terms(raw: str) -> list[str]:
    terms: list[str] = []
    for part in (raw or "").split(","):
        token = part.strip()
        if not token:
            continue
        if token not in terms:
            terms.append(token)
    return terms


def _norm_rule_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def _dynamic_hard_filter_rules(role_result: dict) -> list[str]:
    """Return only JD-specific (non-universal) hard-filter rule texts."""
    hard_filters = role_result.get("hard_filters", []) if isinstance(role_result, dict) else []
    out: list[str] = []
    for item in hard_filters:
        if not isinstance(item, dict):
            continue
        if bool(item.get("universal")):
            continue
        rule = str(item.get("rule", "")).strip()
        if rule and rule not in out:
            out.append(rule)
    return out


def _get_role_session_overrides(role_id: str) -> list[str]:
    all_overrides = session.get("role_filter_overrides", {})
    if not isinstance(all_overrides, dict):
        return []
    role_overrides = all_overrides.get(role_id, [])
    if not isinstance(role_overrides, list):
        return []
    return [str(x).strip() for x in role_overrides if str(x).strip()]


def _set_role_session_overrides(role_id: str, overrides: list[str]) -> None:
    all_overrides = session.get("role_filter_overrides", {})
    if not isinstance(all_overrides, dict):
        all_overrides = {}
    if overrides:
        all_overrides[role_id] = overrides
    elif role_id in all_overrides:
        del all_overrides[role_id]
    session["role_filter_overrides"] = all_overrides


def _get_role_session_removed_keywords(role_id: str) -> list[str]:
    all_removed = session.get("role_removed_keywords", {})
    if not isinstance(all_removed, dict):
        return []
    role_removed = all_removed.get(role_id, [])
    if not isinstance(role_removed, list):
        return []
    return [str(x).strip() for x in role_removed if str(x).strip()]


def _set_role_session_removed_keywords(role_id: str, removed_keywords: list[str]) -> None:
    all_removed = session.get("role_removed_keywords", {})
    if not isinstance(all_removed, dict):
        all_removed = {}
    if removed_keywords:
        all_removed[role_id] = removed_keywords
    elif role_id in all_removed:
        del all_removed[role_id]
    session["role_removed_keywords"] = all_removed


def _get_role_session_disabled_hard_filters(role_id: str) -> list[str]:
    all_disabled = session.get("role_disabled_hard_filters", {})
    if not isinstance(all_disabled, dict):
        return []
    role_disabled = all_disabled.get(role_id, [])
    if not isinstance(role_disabled, list):
        return []
    return [str(x).strip() for x in role_disabled if str(x).strip()]


def _set_role_session_disabled_hard_filters(role_id: str, disabled_filters: list[str]) -> None:
    all_disabled = session.get("role_disabled_hard_filters", {})
    if not isinstance(all_disabled, dict):
        all_disabled = {}
    if disabled_filters:
        all_disabled[role_id] = disabled_filters
    elif role_id in all_disabled:
        del all_disabled[role_id]
    session["role_disabled_hard_filters"] = all_disabled


def _get_role_session_requirements(role_id: str) -> Optional[dict]:
    all_requirements = session.get("role_live_requirements", {})
    if not isinstance(all_requirements, dict):
        return None
    snapshot = all_requirements.get(role_id)
    if not isinstance(snapshot, dict):
        return None
    return snapshot


def _set_role_session_requirements(role_id: str, snapshot: Optional[dict]) -> None:
    all_requirements = session.get("role_live_requirements", {})
    if not isinstance(all_requirements, dict):
        all_requirements = {}
    if snapshot:
        all_requirements[role_id] = snapshot
    elif role_id in all_requirements:
        del all_requirements[role_id]
    session["role_live_requirements"] = all_requirements


def _resolve_anchor_job_id(role: dict, role_result: Optional[dict]) -> Optional[str]:
    role_result = role_result or {}
    anchor_job_id = role_result.get("anchor_job_id")
    if not anchor_job_id:
        all_job_ids = role_result.get("all_job_ids") or []
        anchor_job_id = all_job_ids[0] if all_job_ids else None
    if not anchor_job_id:
        role_job_ids = role.get("job_ids") or []
        anchor_job_id = role_job_ids[0] if role_job_ids else None
    return str(anchor_job_id) if anchor_job_id else None


def _refresh_live_requirements(role: dict, role_result: Optional[dict]) -> Optional[dict]:
    from sa_candidate_finder.config import load_config
    from sa_candidate_finder.pipeline.extractor import extract_constraints
    from sa_candidate_finder.manatal_jobs import strip_html

    jobs = fetch_all_jobs(MANATAL_API_KEY)
    anchor_job_id = _resolve_anchor_job_id(role, role_result)
    if not anchor_job_id:
        return None

    anchor_job = next((j for j in jobs if str(j.get("id")) == anchor_job_id), None)
    if not anchor_job:
        return None

    raw_desc = anchor_job.get("description") or ""
    jd_text = strip_html(raw_desc) if raw_desc else (anchor_job.get("position_name") or "")
    cfg = load_config()
    extraction = extract_constraints(jd_text, cfg)
    return {
        "anchor_job_id": anchor_job_id,
        "llm_keywords": list(extraction.get("keywords", []) or []),
        "hard_filters": list(extraction.get("dealbreakers", []) or []),
        "refreshed_at": datetime.now(timezone.utc).isoformat(),
    }


def _infer_status(jobs: list[dict], role_result: Optional[dict]) -> str:
    if jobs and all(j.get("status") == "on_hold" for j in jobs):
        return "On hold"
    if role_result and role_result.get("top_candidates"):
        return "Shortlist review"
    return "Sourcing"


def _build_roles() -> list[dict]:
    jobs = fetch_all_jobs(MANATAL_API_KEY)
    active_jobs = [j for j in jobs if j.get("status") in ("active", "on_hold")]

    grouped: dict[str, list[dict]] = defaultdict(list)
    for job in active_jobs:
        grouped[_normalize_role(job.get("position_name", ""))].append(job)

    roles: list[dict] = []
    for role_key, role_jobs in grouped.items():
        if not role_key:
            continue
        sample_name = role_jobs[0].get("position_name", "Unnamed role")
        role_id = _role_id_from_name(sample_name)
        role_result = _load_role_results(role_id)
        candidates = int((role_result or {}).get("total_candidates_in_pool", 0))
        full_list = (role_result or {}).get("candidates", [])
        tier_a = sum(1 for c in full_list if str(c.get("tier", "")).upper() == "A")
        status = _infer_status(role_jobs, role_result)

        roles.append(
            {
                "id": role_id,
                "name": sample_name,
                "num_jds": len(role_jobs),
                "candidates": candidates,
                "tier_a": tier_a,
                "status": status,
                "job_ids": [str(j.get("id")) for j in role_jobs],
                "has_results": bool(role_result),
            }
        )

    roles.sort(key=lambda r: (r["candidates"], r["num_jds"]), reverse=True)
    return _enrich_roles(roles)


def _enrich_roles(roles: list[dict]) -> list[dict]:
    out: list[dict] = []
    for role in roles:
        item = dict(role)
        meta = STATUS_META.get(item["status"], {"color": "taupe", "hex": "#CDBCAB"})
        item["status_color"] = meta["color"]
        item["status_hex"] = meta["hex"]
        item["candidates_fmt"] = f"{int(item.get('candidates', 0)):,}"
        out.append(item)
    return out


def _create_rerank_job(role_id: str, session_id: str) -> str:
    job_id = uuid.uuid4().hex
    payload = {
        "id": job_id,
        "role_id": role_id,
        "session_id": session_id,
        "status": "running",
        "progress": 5,
        "message": "Queued rerank job",
        "created_at": int(time.time()),
        "updated_at": int(time.time()),
        "done_state": "",
    }
    with RERANK_LOCK:
        RERANK_JOBS[job_id] = payload
    return job_id


def _update_rerank_job(job_id: str, **fields) -> None:
    with RERANK_LOCK:
        job = RERANK_JOBS.get(job_id)
        if not job:
            return
        job.update(fields)
        job["updated_at"] = int(time.time())


def _run_rerank_job(job_id: str, role_id: str, cmd: list[str], session_results_dir: str) -> None:
    _update_rerank_job(job_id, progress=10, message="Starting rerank process")
    try:
        env = os.environ.copy()
        env["SA_RESULTS_DIR"] = session_results_dir
        src_path = str(ROOT / "src")
        existing_pythonpath = env.get("PYTHONPATH", "")
        env["PYTHONPATH"] = src_path if not existing_pythonpath else f"{src_path}{os.pathsep}{existing_pythonpath}"
        proc = subprocess.Popen(
            cmd,
            cwd=str(ROOT),
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    except Exception as exc:
        app.logger.error("rerank start failed for role %s: %s", role_id, exc)
        _update_rerank_job(job_id, status="failed", progress=100, message="Failed to start rerank", done_state="error")
        return

    progress = 10
    last_line = "Running rerank"
    current_source_job_id = ""
    current_candidate_index = 0
    current_candidate_total = 0
    current_phase = 0  # 0=unknown, 1=phase 1, 2=phase 2, 3=finalizing

    _job_from_cache_line = re.compile(r"job\s+(\d+)", re.IGNORECASE)
    _job_from_matches_line = re.compile(r"/jobs/(\d+)/matches", re.IGNORECASE)
    _candidate_progress_line = re.compile(
        r"job\s+(\d+)\s*:\s*candidate\s+(\d+)\s*/\s*(\d+)",
        re.IGNORECASE,
    )
    try:
        assert proc.stdout is not None
        for raw_line in proc.stdout:
            line = (raw_line or "").strip()
            if not line:
                continue
            last_line = line
            lower = line.lower()

            progress_match = _candidate_progress_line.search(line)
            if progress_match:
                current_source_job_id = progress_match.group(1)
                current_candidate_index = int(progress_match.group(2))
                current_candidate_total = int(progress_match.group(3))
            else:
                job_match = _job_from_cache_line.search(line) or _job_from_matches_line.search(line)
                if job_match:
                    current_source_job_id = job_match.group(1)

            if "phase 1" in lower:
                progress = max(progress, 25)
                current_phase = 1
            elif "phase 2" in lower:
                progress = max(progress, 45)
                current_phase = 2
            elif "running claude evaluation" in lower:
                progress = max(progress, 70)
                current_phase = 2
            elif "after claude evaluation" in lower:
                progress = max(progress, 82)
                current_phase = 3
            elif "top " in lower and "candidates" in lower:
                progress = max(progress, 92)
                current_phase = 3
            elif "wrote role run json" in lower or "saved role run json" in lower:
                progress = max(progress, 96)
                current_phase = 3

            # Advance progress dynamically based on candidate x/y within a phase window
            if current_candidate_total > 0:
                frac = current_candidate_index / current_candidate_total
                if current_phase <= 1:
                    # Phase 1 window: 25 → 43
                    progress = max(progress, int(25 + frac * 18))
                    current_phase = max(current_phase, 1)
                elif current_phase == 2:
                    # Phase 2 window: 45 → 68
                    progress = max(progress, int(45 + frac * 23))
                    current_phase = max(current_phase, 2)

                # Avoid a perceived stall at 68% when phase-2 candidate processing is done
                # but the run is still handling retries/output writes before completion markers appear.
                if current_phase == 2 and current_candidate_index >= current_candidate_total:
                    progress = max(progress, 72)
                    current_phase = 3

            phase_prefix = ""
            if current_phase == 1:
                phase_prefix = "[Phase 1/2: Fetching candidates] "
            elif current_phase == 2:
                phase_prefix = "[Phase 2/2: LLM evaluation] "
            elif current_phase == 3:
                phase_prefix = "[Finalizing] "

            display_message = line
            if current_source_job_id and current_candidate_total > 0:
                display_message = (
                    f"{phase_prefix}Job {current_source_job_id}: candidate {current_candidate_index}/"
                    f"{current_candidate_total} | {line}"
                )
            elif current_source_job_id:
                display_message = f"{phase_prefix}Job {current_source_job_id} | {line}"
            else:
                display_message = f"{phase_prefix}{line}"

            _update_rerank_job(job_id, progress=progress, message=display_message[-220:])
    except Exception as exc:
        app.logger.warning("rerank output read failed for role %s: %s", role_id, exc)

    return_code = proc.wait()
    if return_code == 0:
        publish_message = "Re-rank completed"
        try:
            session_result_path = Path(session_results_dir) / f"role_{role_id}.json"
            global_result_path = _role_results_path(role_id)
            if session_result_path.exists():
                global_result_path.parent.mkdir(parents=True, exist_ok=True)
                tmp_path = global_result_path.parent / f".{global_result_path.name}.{job_id}.tmp"
                tmp_path.write_text(session_result_path.read_text(encoding="utf-8"), encoding="utf-8")
                tmp_path.replace(global_result_path)
                publish_message = "Re-rank completed and published"
            else:
                publish_message = "Re-rank completed (publish skipped: no result file)"
        except Exception as exc:
            app.logger.warning("rerank publish failed for role %s: %s", role_id, exc)
            publish_message = "Re-rank completed (publish failed)"

        _update_rerank_job(job_id, status="completed", progress=100, message=publish_message, done_state="ok")
        return

    app.logger.error("rerank failed for role %s: %s", role_id, last_line)
    _update_rerank_job(job_id, status="failed", progress=100, message=last_line[-140:], done_state="failed")


@app.context_processor
def inject_nav_roles():
    roles = _build_roles()
    nav = sorted(
        [r for r in roles if r.get("status") != "On hold"],
        key=lambda r: r["name"].lower(),
    )
    return {"nav_roles": nav}


@app.route("/")
def index():
    roles = _build_roles()
    return render_template("index.html", roles=roles)


@app.route("/refresh-roles", methods=["POST"])
def refresh_roles():
    try:
        jobs = fetch_all_jobs(MANATAL_API_KEY, force_refresh=True)
        return redirect(url_for("index", refreshed="ok", jobs=str(len(jobs))))
    except Exception:
        return redirect(url_for("index", refreshed="error"))


@app.route("/health")
def health_check():
    """Health check endpoint for monitoring."""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": "1.0.0"
    })


@app.route("/role/<role_id>")
def role_detail(role_id: str):
    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        abort(404)

    role_result = _load_role_results(role_id) or {}
    candidates = role_result.get("candidates", [])

    # Load the most recently synced stage snapshot (local file read — no API call).
    # Use Sync Stages button to refresh from Manatal when candidates have moved pipeline stages.
    live_stages = load_role_stage_snapshot(role_id)

    # Overwrite each candidate's manatal_stage with the synced value so the column is current
    for c in candidates:
        live = live_stages.get(str(c.get("id", "")))
        if live is not None:
            c["manatal_stage"] = live

    # Filter out candidates whose current stage is not in the allowed set
    candidates = [c for c in candidates if not is_excluded_stage(c.get("manatal_stage", ""))]
    # Sort: Tier A → B → C → untiered (disqualified or keyword-only), then by score desc within each group
    _tier_order = {"A": 0, "B": 1, "C": 2}
    candidates = sorted(
        candidates,
        key=lambda c: (
            _tier_order.get(str(c.get("tier", "")).upper(), 3),
            -(c.get("fit_score") or 0),
        ),
    )
    role["top_candidate_count"] = len(role_result.get("top_candidates", []))
    role["llm_evaluated_candidates"] = int(role_result.get("llm_evaluated_candidates", 0))
    role["total_ranked_candidates"] = int(role_result.get("total_ranked_candidates", 0))
    role["total_candidates_in_pool"] = int(role_result.get("total_candidates_in_pool", 0))
    role["tier_b"] = sum(1 for c in candidates if str(c.get("tier", "")).upper() == "B")
    role["last_ran_at"] = role_result.get("ran_at", "")
    last_ran_at = _parse_iso_utc(role["last_ran_at"])
    live_requirements = _get_role_session_requirements(role_id) or {}
    live_requirements_at = _parse_iso_utc(live_requirements.get("refreshed_at", "")) if live_requirements else None
    requirements_payload = role_result
    if live_requirements and live_requirements_at and (not last_ran_at or live_requirements_at >= last_ran_at):
        requirements_payload = live_requirements
        role["requirements_source"] = "live"
    else:
        role["requirements_source"] = "snapshot"
    jobs_refreshed_at = _jobs_cache_fetched_at()
    latest_requirements_at = max(
        [dt for dt in [last_ran_at, live_requirements_at] if dt],
        default=None,
    )
    role["requirements_stale"] = bool(
        latest_requirements_at and jobs_refreshed_at and jobs_refreshed_at > latest_requirements_at
    )
    live_hard_filters = [
        f for f in (requirements_payload.get("hard_filters", []) or [])
        if isinstance(f, dict) and not bool(f.get("universal"))
    ]
    # If a live refresh returned no JD-specific filters, fall back to the
    # results-file filters so a failed or sparse LLM extraction doesn't wipe
    # requirements that the full re-rank successfully extracted.
    if not live_hard_filters and role["requirements_source"] == "live":
        live_hard_filters = [
            f for f in (role_result.get("hard_filters", []) or [])
            if isinstance(f, dict) and not bool(f.get("universal"))
        ]
    role["hard_filters"] = live_hard_filters
    live_keywords = requirements_payload.get("llm_keywords", [])
    if not live_keywords and role["requirements_source"] == "live":
        live_keywords = role_result.get("llm_keywords", [])
    role["llm_keywords"] = live_keywords
    role["user_keywords"] = role_result.get("user_keywords", [])
    role["final_keyword_set"] = role_result.get("final_keyword_set", [])
    role["session_override_filters"] = _get_role_session_overrides(role_id)
    role["session_removed_keywords"] = _get_role_session_removed_keywords(role_id)
    session_disabled = _get_role_session_disabled_hard_filters(role_id)
    allowed_dynamic_rules = {_norm_rule_text(r) for r in _dynamic_hard_filter_rules(role_result)}
    sanitized_disabled = [
        r for r in session_disabled
        if _norm_rule_text(r) in allowed_dynamic_rules
    ]
    if sanitized_disabled != session_disabled:
        _set_role_session_disabled_hard_filters(role_id, sanitized_disabled)
    role["session_disabled_hard_filters"] = sanitized_disabled

    # Attach Goodfit invite status to each candidate
    for c in candidates:
        invite = _load_goodfit_invite(str(c.get("id", "")), role_id)
        c["goodfit_invite"] = invite

    return render_template("role.html", role=role, candidates=candidates)


@app.route("/role/<role_id>/refresh-requirements", methods=["POST"])
def role_refresh_requirements(role_id: str):
    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        abort(404)

    role_result = _load_role_results(role_id) or {}
    snapshot = _refresh_live_requirements(role, role_result)
    if not snapshot:
        return redirect(url_for("role_detail", role_id=role_id, requirements="error"))

    _set_role_session_requirements(role_id, snapshot)
    return redirect(url_for("role_detail", role_id=role_id, requirements="refreshed"))


@app.route("/role/<role_id>/refresh-and-rerank", methods=["POST"])
def role_refresh_and_rerank(role_id: str):
    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        abort(404)

    role_result = _load_role_results(role_id) or {}

    # Refresh live requirements from Manatal and store in session for display
    snapshot = _refresh_live_requirements(role, role_result)
    if snapshot:
        _set_role_session_requirements(role_id, snapshot)

    anchor_job_id = _resolve_anchor_job_id(role, role_result)
    if not anchor_job_id:
        return redirect(url_for("role_detail", role_id=role_id, rerank="no-job-id"))

    for key, msg in [
        (OPENAI_API_KEY, "OPENAI_API_KEY is not configured. Set OPENAI_API_KEY in the environment."),
        (MANATAL_API_KEY, "MANATAL_API_KEY is not configured. Set MANATAL_API_KEY in the environment."),
        (MANATAL_BASE_URL, "MANATAL_BASE_URL is not configured. Set MANATAL_BASE_URL in the environment."),
    ]:
        if not key:
            return redirect(url_for("role_detail", role_id=role_id, rerank="failed", message=msg))

    # Carry forward any existing session overrides
    if "keywords" in request.form:
        _set_role_session_overrides(role_id, _parse_comma_terms(request.form.get("keywords") or ""))
    if "removed_keywords" in request.form:
        _set_role_session_removed_keywords(role_id, _parse_comma_terms(request.form.get("removed_keywords") or ""))
    dynamic_rule_norm = {_norm_rule_text(r) for r in _dynamic_hard_filter_rules(role_result)}
    if "disabled_hard_filters" in request.form:
        disabled = [
            r for r in _parse_comma_terms(request.form.get("disabled_hard_filters") or "")
            if _norm_rule_text(r) in dynamic_rule_norm
        ]
        _set_role_session_disabled_hard_filters(role_id, disabled)

    session_override_filters = _get_role_session_overrides(role_id)
    session_removed_keywords = _get_role_session_removed_keywords(role_id)
    session_disabled_hard_filters = [
        r for r in _get_role_session_disabled_hard_filters(role_id)
        if _norm_rule_text(r) in dynamic_rule_norm
    ]
    _set_role_session_disabled_hard_filters(role_id, session_disabled_hard_filters)

    cmd = [
        sys.executable,
        str(ROOT / "src" / "sa_candidate_finder" / "cli.py"),
        "agentic-search",
        "--job-id",
        str(anchor_job_id),
        "--rerank-fast",
    ]
    keyword_directives: list[str] = []
    if session_override_filters:
        keyword_directives.extend(session_override_filters)
    if session_removed_keywords:
        keyword_directives.extend([f"-{kw}" for kw in session_removed_keywords])
    if keyword_directives:
        cmd.extend(["--keywords", ",".join(keyword_directives)])
    if session_disabled_hard_filters:
        cmd.extend(["--disable-hard-filter-rules", "||".join(session_disabled_hard_filters)])

    session_id = _get_session_id(create=True) or uuid.uuid4().hex
    session_results_dir = _session_results_dir(session_id)
    session_results_dir.mkdir(parents=True, exist_ok=True)

    job_id = _create_rerank_job(role_id, session_id)
    threading.Thread(
        target=_run_rerank_job,
        args=(job_id, role_id, cmd, str(session_results_dir)),
        daemon=True,
    ).start()
    return redirect(url_for("role_detail", role_id=role_id, rerank="running", job=job_id))


@app.route("/role/<role_id>/download-excel")
def role_download_excel(role_id: str):
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    role_result = _load_role_results(role_id)
    if not role_result:
        abort(404)

    all_candidates = role_result.get("candidates", [])
    _tier_order = {"A": 0, "B": 1}
    tier_ab = sorted(
        [c for c in all_candidates if str(c.get("tier", "")).upper() in ("A", "B")],
        key=lambda c: (
            _tier_order.get(str(c.get("tier", "")).upper(), 2),
            -(c.get("fit_score") or 0),
        ),
    )
    if not tier_ab:
        abort(404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tier A & B Candidates"

    headers = [
        "Rank", "Name", "Tier", "Fit Score", "Current Position",
        "Current Company", "Location", "Manatal Stage",
        "Matched Keywords", "Strengths", "Risks", "Rationale",
    ]
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="5D1C34", end_color="5D1C34", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    tier_a_fill = PatternFill(start_color="FFF8EC", end_color="FFF8EC", fill_type="solid")
    tier_b_fill = PatternFill(start_color="F4F7F3", end_color="F4F7F3", fill_type="solid")

    for row_idx, c in enumerate(tier_ab, 2):
        tier = str(c.get("tier") or "").upper()
        row_fill = tier_a_fill if tier == "A" else tier_b_fill
        row_data = [
            c.get("rank") or (row_idx - 1),
            c.get("name") or "",
            tier,
            c.get("fit_score"),
            c.get("current_position") or "",
            c.get("current_company") or "",
            c.get("location") or "",
            c.get("manatal_stage") or "",
            ", ".join(c.get("matched_keywords") or []),
            "; ".join(c.get("strengths") or []),
            "; ".join(c.get("risks") or []),
            c.get("rationale") or "",
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [7, 24, 6, 9, 28, 28, 20, 18, 38, 48, 48, 70]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{role_id}_Tier_AB_Candidates.xlsx"
    return Response(
        buf.read(),
        headers={
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@app.route("/role/<role_id>/rerank", methods=["GET", "POST"])
def role_rerank(role_id: str):
    if request.method == "GET":
        return redirect(url_for("role_detail", role_id=role_id))

    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        abort(404)

    role_result = _load_role_results(role_id) or {}
    anchor_job_id = _resolve_anchor_job_id(role, role_result)
    if not anchor_job_id:
        return redirect(url_for("role_detail", role_id=role_id, rerank="no-job-id"))

    if not OPENAI_API_KEY:
        return redirect(
            url_for(
                "role_detail",
                role_id=role_id,
                rerank="failed",
                message="OPENAI_API_KEY is not configured. Set OPENAI_API_KEY in the environment.",
            )
        )
    if not MANATAL_API_KEY:
        return redirect(
            url_for(
                "role_detail",
                role_id=role_id,
                rerank="failed",
                message="MANATAL_API_KEY is not configured. Set MANATAL_API_KEY in the environment.",
            )
        )
    if not MANATAL_BASE_URL:
        return redirect(
            url_for(
                "role_detail",
                role_id=role_id,
                rerank="failed",
                message="MANATAL_BASE_URL is not configured. Set MANATAL_BASE_URL in the environment.",
            )
        )

    if "keywords" in request.form:
        session_terms = _parse_comma_terms(request.form.get("keywords") or "")
        _set_role_session_overrides(role_id, session_terms)
    if "removed_keywords" in request.form:
        removed_terms = _parse_comma_terms(request.form.get("removed_keywords") or "")
        _set_role_session_removed_keywords(role_id, removed_terms)
    dynamic_rule_norm = {
        _norm_rule_text(r) for r in _dynamic_hard_filter_rules(role_result)
    }
    if "disabled_hard_filters" in request.form:
        disabled_filters = _parse_comma_terms(request.form.get("disabled_hard_filters") or "")
        disabled_filters = [
            r for r in disabled_filters
            if _norm_rule_text(r) in dynamic_rule_norm
        ]
        _set_role_session_disabled_hard_filters(role_id, disabled_filters)

    session_override_filters = _get_role_session_overrides(role_id)
    session_removed_keywords = _get_role_session_removed_keywords(role_id)
    session_disabled_hard_filters = [
        r for r in _get_role_session_disabled_hard_filters(role_id)
        if _norm_rule_text(r) in dynamic_rule_norm
    ]
    _set_role_session_disabled_hard_filters(role_id, session_disabled_hard_filters)

    cmd = [
        sys.executable,
        str(ROOT / "src" / "sa_candidate_finder" / "cli.py"),
        "agentic-search",
        "--job-id",
        str(anchor_job_id),
        "--rerank-fast",
    ]
    keyword_directives: list[str] = []
    if session_override_filters:
        keyword_directives.extend(session_override_filters)
    if session_removed_keywords:
        keyword_directives.extend([f"-{kw}" for kw in session_removed_keywords])
    if keyword_directives:
        cmd.extend(["--keywords", ",".join(keyword_directives)])
    if session_disabled_hard_filters:
        cmd.extend(["--disable-hard-filter-rules", "||".join(session_disabled_hard_filters)])

    session_id = _get_session_id(create=True) or uuid.uuid4().hex
    session_results_dir = _session_results_dir(session_id)
    session_results_dir.mkdir(parents=True, exist_ok=True)

    job_id = _create_rerank_job(role_id, session_id)
    worker = threading.Thread(
        target=_run_rerank_job,
        args=(job_id, role_id, cmd, str(session_results_dir)),
        daemon=True,
    )
    worker.start()
    return redirect(url_for("role_detail", role_id=role_id, rerank="running", job=job_id))


@app.route("/role/<role_id>/rerank-status/<job_id>")
def role_rerank_status(role_id: str, job_id: str):
    with RERANK_LOCK:
        job = dict(RERANK_JOBS.get(job_id) or {})

    session_id = _get_session_id(create=False)
    if (
        not job
        or str(job.get("role_id")) != str(role_id)
        or str(job.get("session_id", "")) != str(session_id or "")
    ):
        return jsonify({"ok": False, "error": "job-not-found"}), 404

    return jsonify(
        {
            "ok": True,
            "job_id": job_id,
            "status": job.get("status", "unknown"),
            "progress": int(job.get("progress", 0)),
            "message": job.get("message", ""),
            "done_state": job.get("done_state", ""),
        }
    )


@app.route("/role/<role_id>/clear-overrides")
def role_clear_overrides(role_id: str):
    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        abort(404)
    _set_role_session_overrides(role_id, [])
    _set_role_session_removed_keywords(role_id, [])
    _set_role_session_disabled_hard_filters(role_id, [])

    # Also clear the role result snapshot for this browser session so stale
    # keyword sets (from prior overrides) are not displayed after reset.
    sid = _get_session_id(create=False)
    if sid:
        try:
            session_result_path = _role_results_path(role_id, sid)
            if session_result_path.exists():
                session_result_path.unlink()
        except Exception:
            pass

    return redirect(url_for("role_detail", role_id=role_id, rerank="overrides-cleared"))


@app.route("/role/<role_id>/candidate/<candidate_id>")
def candidate_detail(role_id: str, candidate_id: str):
    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        abort(404)

    role_result = _load_role_results(role_id) or {}
    candidate = next(
        (c for c in role_result.get("candidates", []) if str(c.get("id")) == str(candidate_id)),
        None,
    )
    if not candidate:
        abort(404)

    goodfit_invite = _load_goodfit_invite(str(candidate_id), role_id)

    return render_template(
        "candidate.html",
        role=role,
        c=candidate,
        hard_filters=role_result.get("hard_filters", []),
        llm_eval_policy=role_result.get("llm_eval_policy", {}),
        final_keyword_set=role_result.get("final_keyword_set", []),
        goodfit_invite=goodfit_invite,
    )


# ---------------------------------------------------------------------------
# Stage sync (refresh Manatal pipeline stages without a full re-rank)
# ---------------------------------------------------------------------------

@app.route("/role/<role_id>/sync-stages", methods=["POST"])
def sync_stages(role_id: str):
    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        return jsonify({"success": False, "error": "Role not found"}), 404

    try:
        stages = sync_role_stages(MANATAL_API_KEY, role_id, role.get("job_ids", []))
        return jsonify({"success": True, "candidates_synced": len(stages)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Goodfit AI interview helpers
# ---------------------------------------------------------------------------

def _goodfit_invite_path(candidate_id: str) -> Path:
    return ROOT / "cache" / "candidates" / f"candidate_{candidate_id}.json"


def _load_goodfit_invite(candidate_id: str, role_id: str) -> Optional[dict]:
    path = _goodfit_invite_path(candidate_id)
    try:
        if not path.exists():
            return None
        cached = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(cached, dict):
            return None
        return cached.get("goodfit_invites", {}).get(role_id)
    except Exception:
        return None


def _save_goodfit_invite(candidate_id: str, role_id: str, invite_data: dict) -> None:
    path = _goodfit_invite_path(candidate_id)
    try:
        if path.exists():
            cached = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(cached, dict):
                cached = {}
        else:
            cached = {}
        invites = cached.get("goodfit_invites", {})
        if not isinstance(invites, dict):
            invites = {}
        invites[role_id] = invite_data
        cached["goodfit_invites"] = invites
        path.write_text(json.dumps(cached), encoding="utf-8")
    except Exception as e:
        print(f"[Goodfit] Failed to persist invite data for candidate {candidate_id}: {e}", flush=True)


def _delete_goodfit_invite(candidate_id: str, role_id: str) -> None:
    path = _goodfit_invite_path(candidate_id)
    try:
        if not path.exists():
            return
        cached = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(cached, dict):
            return
        invites = cached.get("goodfit_invites", {})
        if isinstance(invites, dict):
            invites.pop(role_id, None)
            cached["goodfit_invites"] = invites
            path.write_text(json.dumps(cached), encoding="utf-8")
    except Exception as e:
        print(f"[Goodfit] Failed to delete invite data for candidate {candidate_id}: {e}", flush=True)


@app.route("/role/<role_id>/candidate/<candidate_id>/send-goodfit-interview", methods=["POST"])
def send_goodfit_interview(role_id: str, candidate_id: str):
    """Send a Goodfit AI interview invite to a candidate and move them in Manatal."""
    import httpx as _httpx
    from sa_candidate_finder import goodfit as _goodfit
    from sa_candidate_finder.manatal_candidates import fetch_candidate_contact, update_match_stage

    if not GOODFIT_API_KEY:
        return jsonify({"success": False, "error": "GOODFIT_API_KEY is not configured on the server"}), 500

    roles = _build_roles()
    role = next((r for r in roles if r["id"] == role_id), None)
    if not role:
        return jsonify({"success": False, "error": "Role not found"}), 404

    # Check if already sent
    existing = _load_goodfit_invite(candidate_id, role_id)
    force_resend = request.args.get("resend") == "1"
    if existing and existing.get("application_id"):
        app_id = existing["application_id"]
        # Verify the application still exists in Goodfit (could have been deleted).
        if _goodfit.application_exists(GOODFIT_API_KEY, app_id):
            cached_url = existing.get("direct_apply_url", "")
            if force_resend:
                return jsonify({
                    "success": True,
                    "resent": True,
                    "application_id": app_id,
                    "goodfit_job_title": existing.get("goodfit_job_title", ""),
                    "direct_apply_url": cached_url,
                })
            return jsonify({
                "success": True,
                "already_sent": True,
                "application_id": app_id,
                "goodfit_job_title": existing.get("goodfit_job_title", ""),
                "direct_apply_url": cached_url,
            })
        # Application no longer exists in Goodfit — clear stale cache and re-invite.
        print(f"[Goodfit] Application {app_id} not found in Goodfit; clearing stale invite and re-sending for candidate {candidate_id}.", flush=True)
        _delete_goodfit_invite(candidate_id, role_id)

    # Get candidate name + email — try cache first, then Manatal API
    cache_path = ROOT / "cache" / "candidates" / f"candidate_{candidate_id}.json"
    candidate_name = ""
    candidate_email = ""
    try:
        if cache_path.exists():
            raw = json.loads(cache_path.read_text(encoding="utf-8"))
            data = raw.get("data", {}) if isinstance(raw, dict) else {}
            candidate_name = data.get("name") or data.get("full_name", "")
            candidate_email = data.get("email", "")
    except Exception:
        pass

    if not candidate_email:
        contact = fetch_candidate_contact(MANATAL_API_KEY, candidate_id)
        candidate_email = contact.get("email", "")
        if not candidate_name:
            candidate_name = contact.get("name", "")

    if not candidate_email:
        return jsonify({"success": False, "error": "No email address found for this candidate in Manatal"}), 400

    # Find the matching Goodfit job by role title
    try:
        goodfit_job = _goodfit.find_goodfit_job_by_title(GOODFIT_API_KEY, role["name"])
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to fetch Goodfit jobs: {e}"}), 500

    if not goodfit_job:
        return jsonify({
            "success": False,
            "error": f"No active Goodfit role found matching '{role['name']}'. "
                     "Please create a matching role in Goodfit first.",
        }), 404

    # Send the invite
    try:
        app_data = _goodfit.send_interview_invite(
            GOODFIT_API_KEY,
            goodfit_job["id"],
            candidate_name,
            candidate_email,
        )
    except Exception as e:
        return jsonify({"success": False, "error": f"Goodfit API error: {e}"}), 500

    # Extract application ID from the response (handles different response shapes)
    application_id = (
        app_data.get("id")
        or (app_data.get("application") or {}).get("id")
        or ""
    )

    # Build the direct apply URL — candidates use this instead of the magic link email,
    # which is broken due to Goodfit's API creating UUID-based users that can't
    # authenticate via the platform's Firebase auth.
    job_slug = goodfit_job.get("slug", "")
    from sa_candidate_finder.goodfit import GOODFIT_APPLY_BASE
    direct_apply_url = f"{GOODFIT_APPLY_BASE}/{job_slug}" if job_slug else ""

    # Persist invite in the candidate's local cache
    _save_goodfit_invite(candidate_id, role_id, {
        "application_id": str(application_id),
        "goodfit_job_id": goodfit_job["id"],
        "goodfit_job_title": goodfit_job.get("title", ""),
        "goodfit_job_slug": job_slug,
        "direct_apply_url": direct_apply_url,
        "candidate_email": candidate_email,
        "sent_at": time.time(),
    })

    # Update Manatal stage (best-effort — failure does not block the response)
    for jid in role.get("job_ids", []):
        try:
            if update_match_stage(MANATAL_API_KEY, jid, candidate_id, "Goodfit Interview Sent"):
                break
        except Exception as e:
            print(f"[Goodfit] Manatal stage update failed for job {jid}: {e}", flush=True)

    # Update the local stage snapshot immediately so the candidate is excluded on the next
    # page load without needing a full Sync Stages call.
    update_stage_snapshot_entry(role_id, candidate_id, "Goodfit Interview Sent")

    return jsonify({
        "success": True,
        "application_id": str(application_id),
        "goodfit_job_title": goodfit_job.get("title", ""),
        "direct_apply_url": direct_apply_url,
    })


@app.route("/role/<role_id>/candidate/<candidate_id>/goodfit-status")
def goodfit_interview_status(role_id: str, candidate_id: str):
    """Return the stored Goodfit invite status for a candidate."""
    invite = _load_goodfit_invite(candidate_id, role_id)
    if not invite:
        return jsonify({"sent": False})
    return jsonify({
        "sent": True,
        "application_id": invite.get("application_id", ""),
        "goodfit_job_title": invite.get("goodfit_job_title", ""),
        "sent_at": invite.get("sent_at", 0),
    })


@app.route("/resume-proxy")
def resume_proxy():
    """Proxy a Manatal signed PDF URL and serve it inline so the browser embeds it."""
    url = request.args.get("url", "")
    if not url or not url.startswith("https://media-assets.manatal.com/"):
        abort(400)
    return _serve_pdf_inline(url)


def _extract_resume_urls(value) -> list[str]:
    """Extract candidate resume URLs from nested Manatal payloads."""
    out: list[str] = []

    def _walk(node):
        if isinstance(node, str):
            v = node.strip()
            if v.startswith("https://media-assets.manatal.com/"):
                out.append(v)
            return
        if isinstance(node, dict):
            for child in node.values():
                _walk(child)
            return
        if isinstance(node, list):
            for child in node:
                _walk(child)

    _walk(value)

    uniq: list[str] = []
    seen: set[str] = set()
    for url in out:
        if url not in seen:
            seen.add(url)
            uniq.append(url)
    return uniq


def _resume_url_expiry_epoch(url: str) -> int:
    """Return signed URL expiry epoch seconds; large future value if no expiry token."""
    try:
        parsed = _urlparse.urlparse(url)
        expires_vals = _urlparse.parse_qs(parsed.query).get("Expires", [])
        if not expires_vals:
            return 2_147_483_647
        return int(expires_vals[0])
    except Exception:
        return 0


def _resume_url_ext(url: str) -> str:
    """Return normalized lowercase file extension (without dot) from a signed resume URL."""
    try:
        path = _urlparse.urlparse(url).path or ""
        tail = path.rsplit("/", 1)[-1]
        if "." not in tail:
            return ""
        return tail.rsplit(".", 1)[-1].lower().strip()
    except Exception:
        return ""


def _pdf_inline_response(pdf_bytes: bytes, *, max_age: int) -> Response:
    """Build an inline PDF response with optional byte-range support for browser viewers."""
    total = len(pdf_bytes)
    base_headers = {
        "Content-Type": "application/pdf",
        "Content-Disposition": 'inline; filename="resume.pdf"',
        "Cache-Control": f"private, max-age={max_age}",
        "X-Content-Type-Options": "nosniff",
        "Accept-Ranges": "bytes",
    }

    range_header = request.headers.get("Range", "")
    if not range_header.startswith("bytes="):
        headers = {**base_headers, "Content-Length": str(total)}
        return Response(pdf_bytes, status=200, headers=headers)

    m = re.match(r"bytes=(\d*)-(\d*)$", range_header.strip())
    if not m:
        headers = {**base_headers, "Content-Range": f"bytes */{total}"}
        return Response(status=416, headers=headers)

    start_s, end_s = m.group(1), m.group(2)
    if not start_s and not end_s:
        headers = {**base_headers, "Content-Range": f"bytes */{total}"}
        return Response(status=416, headers=headers)

    if start_s:
        start = int(start_s)
        end = int(end_s) if end_s else total - 1
    else:
        suffix_len = int(end_s)
        if suffix_len <= 0:
            headers = {**base_headers, "Content-Range": f"bytes */{total}"}
            return Response(status=416, headers=headers)
        start = max(total - suffix_len, 0)
        end = total - 1

    if start >= total or end < start:
        headers = {**base_headers, "Content-Range": f"bytes */{total}"}
        return Response(status=416, headers=headers)

    end = min(end, total - 1)
    chunk = pdf_bytes[start : end + 1]
    headers = {
        **base_headers,
        "Content-Range": f"bytes {start}-{end}/{total}",
        "Content-Length": str(len(chunk)),
    }
    return Response(chunk, status=206, headers=headers)


def _serve_pdf_inline_try(url: str) -> Optional[Response]:
    """Try to fetch and serve PDF for a signed URL; return None when URL is invalid/expired/not a PDF."""
    try:
        req = _urlrequest.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = _urlrequest.urlopen(req, timeout=20)
        data = resp.read()
        if not data.startswith(b"%PDF-"):
            app.logger.warning("resume fetch for %s returned non-PDF content (first bytes: %s)", url[:80], data[:8])
            return None
        return _pdf_inline_response(data, max_age=3600)
    except _urlrequest.HTTPError as _e:
        app.logger.warning("resume PDF fetch HTTP error %s for %s", _e.code, url[:80])
        return None
    except Exception as _e:
        app.logger.error("resume PDF fetch error for %s: %s", url[:80], _e)
        return None


def _serve_pdf_inline(url: str) -> Response:
    try:
        req = _urlrequest.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = _urlrequest.urlopen(req, timeout=20)
        data = resp.read()
    except _urlrequest.HTTPError as _e:
        app.logger.warning("resume PDF fetch HTTP error %s for %s", _e.code, url[:80])
        # 403 = signed URL expired; surface as 404 so the browser shows a clean message
        abort(404 if _e.code in (403, 410) else 502)
    except Exception as _e:
        app.logger.error("resume PDF fetch error for %s: %s", url[:80], _e)
        abort(502)
    if not data.startswith(b"%PDF-"):
        abort(404)
    return _pdf_inline_response(data, max_age=3600)


@app.route("/resume-fresh/<int:candidate_id>")
def resume_fresh(candidate_id: int):
    """Fetch a fresh signed resume URL from the Manatal API and serve it inline.
    Falls back to the cached URL if the API is rate-limited (429)."""
    candidate_urls: list[str] = []

    def _add_urls(urls: list[str]) -> None:
        for u in urls:
            if u and u not in candidate_urls:
                candidate_urls.append(u)

    # --- Check local PDF disk cache first (no API call needed) ---
    try:
        from sa_candidate_finder.manatal_candidates import load_cached_resume_pdf
        cached_pdf = load_cached_resume_pdf(candidate_id)
        if cached_pdf:
            return _pdf_inline_response(cached_pdf, max_age=86400)
    except Exception as _ce:
        app.logger.warning("resume-fresh local cache read failed for %s: %s", candidate_id, _ce)

    # --- Try Manatal API for a fresh signed URL (with retry on 429) ---
    api_url = f"https://api.manatal.com/open/v3/candidates/{candidate_id}/"
    for _attempt in range(3):
        try:
            req = _urlrequest.Request(
                api_url,
                headers={
                    "Authorization": f"Token {MANATAL_API_KEY}",
                    "accept": "application/json",
                    "User-Agent": "Mozilla/5.0",
                },
            )
            resp = _urlrequest.urlopen(req, timeout=20)
            candidate_data = json.loads(resp.read().decode("utf-8"))
            _add_urls(_extract_resume_urls(candidate_data))
            break
        except _urlrequest.HTTPError as _e:
            if _e.code == 429 and _attempt < 2:
                time.sleep(3 * (_attempt + 1))
                continue
            app.logger.warning("resume-fresh API error for candidate %s: %s — trying cache fallback", candidate_id, _e)
            break
        except Exception as _e:
            app.logger.warning("resume-fresh API error for candidate %s: %s — trying cache fallback", candidate_id, _e)
            break

    # --- Fall back to URL stored in candidate cache if API failed / rate-limited ---
    cache_file = ROOT / "cache" / "candidates" / f"candidate_{candidate_id}.json"
    try:
        cached = json.loads(cache_file.read_text(encoding="utf-8"))
        data = cached.get("data", {})
        _add_urls(_extract_resume_urls(data))
    except Exception as _ce:
        app.logger.error("resume-fresh cache fallback failed for candidate %s: %s", candidate_id, _ce)

    # --- Last resort: scan results JSONs for this candidate's resume_url ---
    results_dir = ROOT / "results"
    try:
        for results_file in results_dir.glob("role_*.json"):
            try:
                role_data = json.loads(results_file.read_text(encoding="utf-8"))
                for c in role_data.get("candidates", []):
                    if str(c.get("id")) == str(candidate_id):
                        _add_urls(_extract_resume_urls(c))
            except Exception:
                continue
    except Exception as _re:
        app.logger.error("resume-fresh results fallback failed for candidate %s: %s", candidate_id, _re)

    if not candidate_urls:
        abort(404, description="Resume unavailable: no resume URL found for this candidate.")

    now_epoch = int(time.time())
    ordered_urls = sorted(
        candidate_urls,
        key=lambda u: (_resume_url_expiry_epoch(u) >= now_epoch, _resume_url_expiry_epoch(u)),
        reverse=True,
    )

    for url in ordered_urls:
        ext = _resume_url_ext(url)
        # Some candidates have non-PDF resumes (e.g., .docx). For those, redirect
        # to the fresh signed URL instead of forcing PDF-only embed handling.
        if ext in {"doc", "docx", "rtf", "odt", "txt"}:
            return redirect(url)

        response = _serve_pdf_inline_try(url)
        if response is not None:
            # Save to local disk cache for future requests
            try:
                from sa_candidate_finder.manatal_candidates import save_resume_pdf
                save_resume_pdf(candidate_id, response.get_data())
            except Exception:
                pass
            return response

    abort(
        404,
        description=(
            "Resume temporarily unavailable: all known signed resume links are expired, inaccessible, "
            "or provided in an unsupported format for inline PDF preview. "
            "This usually happens when Manatal API is rate-limited and a fresh link cannot be fetched."
        ),
    )


@app.route("/resume-download/<int:candidate_id>")
def resume_download(candidate_id: int):
    """Download candidate resume as attachment when PDF; otherwise redirect to source URL."""
    response = resume_fresh(candidate_id)

    # Non-PDF resumes are redirected by resume_fresh to a signed source URL.
    if 300 <= response.status_code < 400:
        return response

    pdf_bytes = response.get_data()
    headers = {
        "Content-Type": "application/pdf",
        "Content-Disposition": f'attachment; filename="candidate_{candidate_id}_resume.pdf"',
        "Content-Length": str(len(pdf_bytes)),
        "Cache-Control": "private, max-age=3600",
        "X-Content-Type-Options": "nosniff",
    }
    return Response(pdf_bytes, status=200, headers=headers)


@app.route("/resume-view/<int:candidate_id>")
def resume_view(candidate_id: int):
    """Serve a minimal HTML page that embeds the PDF — prevents browser download dialogs."""
    proxy_url = _urlparse.quote(f"/resume-fresh/{candidate_id}", safe="/:")
    html = f"""<!doctype html>
<html><head><meta charset="utf-8">
<style>html,body{{margin:0;padding:0;height:100%;overflow:hidden}}
embed{{display:block;width:100%;height:100%;border:none}}</style>
</head><body>
<embed src="{proxy_url}" type="application/pdf" width="100%" height="100%">
</body></html>"""
    return Response(html, status=200, headers={"Content-Type": "text/html; charset=utf-8"})


@app.route("/api/roles")
def api_roles():
    return jsonify(_build_roles())


@app.route("/api/jobs")
def api_jobs():
    try:
        jobs = fetch_all_jobs(MANATAL_API_KEY)
        return jsonify({"ok": True, "count": len(jobs), "jobs": jobs[:50]})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


_CHAT_SYSTEM = """\
You are a recruiting assistant helping a recruiter use SeamlessAssist, a candidate-search tool.
You have full context about the current role and (optionally) a specific candidate.

Your job is to:
1. Answer questions about the candidate or role honestly and concisely.
2. Explain evaluation results in plain language.
3. Accept feedback and translate it into actionable advice — e.g. suggest better keywords,
    flag a prompt bias, or recommend a setting change.
4. If the recruiter gives feedback that should change how future searches work, clearly state
    what they should do: tweak keywords, adjust a hard filter, change a threshold, or re-run.

Tone: professional, direct, no fluff. Keep replies under 150 words unless detail is essential.
When suggesting keyword changes, output them as a bullet list prefixed with "Keywords to add:" or "Keywords to remove:".
When suggesting a setting change, prefix the line with "Setting change:".
"""


@app.route("/api/chat", methods=["POST"])
def api_chat():
        body = request.get_json(silent=True) or {}
        role_id = body.get("role_id", "")
        candidate_id = body.get("candidate_id")  # optional
        history = body.get("history", [])  # list of {role, content}
        user_message = body.get("message", "").strip()
        if not user_message:
            return jsonify({"error": "message required"}), 400

        # Build context block
        ctx_lines: list[str] = []
        role_result = _load_role_results(role_id) if role_id else {}
        if role_result:
            ctx_lines.append(f"Role: {role_result.get('role_name', role_id)}")
            ctx_lines.append(f"Keywords active: {', '.join(role_result.get('final_keyword_set', []))}")
            ctx_lines.append(f"Hard filters: {'; '.join(f['rule'] for f in role_result.get('hard_filters', []))}")
            ctx_lines.append(f"Pool size: {role_result.get('total_candidates_in_pool', '?')} candidates")
            ctx_lines.append(f"LLM evaluated: {role_result.get('llm_evaluated_candidates', '?')}")
            ctx_lines.append(f"LLM qualified: {role_result.get('llm_qualified_candidates', '?')}")

        if candidate_id:
            candidates = role_result.get("candidates", [])
            cand = next((c for c in candidates if str(c.get("id")) == str(candidate_id)), None)
            if cand:
                ctx_lines.append("")
                ctx_lines.append(f"Candidate: {cand.get('name')} | {cand.get('current_position')} at {cand.get('current_company')} | {cand.get('location')}")
                ctx_lines.append(f"Tier: {cand.get('tier') or 'Not evaluated'} | Score: {cand.get('fit_score') or 'N/A'}/10")
                ctx_lines.append(f"LLM reviewed: {cand.get('llm_reviewed')} | Qualified: {cand.get('llm_qualified')}")
                if cand.get("disqualify_reason"):
                    ctx_lines.append(f"Disqualify reason: {cand['disqualify_reason']}")
                if cand.get("rationale"):
                    ctx_lines.append(f"Rationale: {cand['rationale']}")
                if cand.get("strengths"):
                    ctx_lines.append(f"Strengths: {'; '.join(cand['strengths'])}")
                if cand.get("risks"):
                    ctx_lines.append(f"Risks: {'; '.join(cand['risks'])}")

        context_block = "\n".join(ctx_lines)
        system_msg = _CHAT_SYSTEM + ("\n\n--- CONTEXT ---\n" + context_block if context_block else "")

        messages = [{"role": "system", "content": system_msg}]
        # Include prior turns (cap at last 10 to stay within token budget)
        for turn in history[-10:]:
            if turn.get("role") in ("user", "assistant") and turn.get("content"):
                messages.append({"role": turn["role"], "content": turn["content"]})
        messages.append({"role": "user", "content": user_message})

        try:
            resp = _openai_client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=messages,
                max_tokens=400,
                temperature=0.4,
            )
            reply = resp.choices[0].message.content.strip()
        except Exception as exc:
            app.logger.error("chat error: %s", exc)
            return jsonify({"error": "AI unavailable"}), 502

        return jsonify({"reply": reply})


if __name__ == "__main__":
    # For development only - production uses Gunicorn
    app.run(debug=True, host="0.0.0.0", port=int(os.getenv("PORT", 5050)))
